import os
import time
from _datetime import datetime
from tkinter import messagebox
from tkinter import *
from tkinter import filedialog

import openpyxl
import openpyxl.utils.exceptions


def get_worksheet():
    root.path = filedialog.askopenfilename\
        (initialdir=desktop_path, title="Выберите таблицу", filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    if not root.path:
        quit_gui()
    try:
        workbook = openpyxl.load_workbook(root.path)
        return workbook.active
    except Exception as error:
        messagebox.showerror("ERROR:", error)
        quit_gui()


def get_cell_by_name(name, mode):
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    if mode != "row" and mode != "col":
        print("Invalid mode ", mode)
        return
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell and re.match(name, str(cell.value), flags=re.UNICODE):
                if mode == "col":
                    return col
                if mode == "row":
                    return row
    messagebox.showerror("Ошибка!", "В таблице нет поля \"" + name + "\"")
    quit_gui()


def get_table() -> {}:
    table = {}

    start_row = get_cell_by_name("Заключенные сделки", "row")
    end_row = get_cell_by_name("ИТОГО", "row")  # todo normal way to find out start and end row
    date_time_col = get_cell_by_name("Дата и время", "col")
    stock_code_col = get_cell_by_name("Код бумаги", "col")
    deal_nbr_col = get_cell_by_name("Номер сделки", "col")
    price_col = get_cell_by_name("Цена", "col")
    operation_col = get_cell_by_name("Направление", "col")
    stocks_nbr_col = get_cell_by_name("Кол-во ЦБ", "col")

    for i in range(start_row + 2, end_row):
        stock = worksheet.cell(row=i, column=stock_code_col).value
        operation = worksheet.cell(row=i, column=operation_col).value
        stocks_nbr = worksheet.cell(row=i, column=stocks_nbr_col).value
        deal_nbr = worksheet.cell(row=i, column=deal_nbr_col).value
        price = worksheet.cell(row=i, column=price_col).value
        date_time_str = worksheet.cell(row=i, column=date_time_col).value
        date_time_str = str(date_time_str).strip()
        timestamp = time.mktime(datetime.strptime(date_time_str, "%Y-%m-%d %H:%M:%S").timetuple())
        if re.match("Продажа", operation, re.UNICODE):
            stocks_nbr = -stocks_nbr
        if stock not in table:
            table[stock] = {"balance": stocks_nbr, "deals": {}}
        else:
            table[stock]["balance"] += stocks_nbr
        if re.match("Купля", operation, re.UNICODE):
            table[stock]["deals"][deal_nbr] = [timestamp, stocks_nbr, price]
    return table


def parse_balances(table):
    for stock_name, stock_info in table.items():
        new_list = list()
        date_list = list(stock_info["deals"].values())
        date_list.reverse()
        balance = stock_info["balance"]
        sum_balance = 0
        for i in range(len(date_list)):
            if sum_balance >= balance:
                break
            sum_balance += date_list[i][1]
            new_list.append([datetime.utcfromtimestamp(date_list[i][0]).strftime("%d.%m.%Y"), date_list[i][1], date_list[i][2]])
        if new_list:
            new_list[-1][1] -= abs(sum_balance - balance)
        new_list.reverse()
        stock_info["deals"] = new_list


def save_table(table):
    root.path = filedialog.asksaveasfilename\
        (initialdir=desktop_path, initialfile="untitled.xlsx", title="Сохранить результат", filetypes=[("Excel files", "*.xlsx")])
    if not root.path:
        quit_gui()
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(["Ценная бумага", "Общий баланс", "Дата покупки", "Объем покупки", "Цена"])

    for stock_name, stock_info in table.items():
        if not stock_info["deals"]:
            continue
        new_ws.append([stock_name, stock_info["balance"]])
        for deal in stock_info["deals"]:
            new_ws.append(["", "", deal[0], deal[1], deal[2]])
        new_ws.append([""])

    new_wb.save(root.path)


def setup_gui():
    root.title("Полезная программа от Олега")
    window_width = root.winfo_reqwidth()
    window_height = root.winfo_reqheight()
    position_right = int(root.winfo_screenwidth() / 2 - window_width / 2)
    position_down = int(root.winfo_screenheight() / 2 - window_height / 2)
    root.geometry("+{}+{}".format(position_right, position_down))


def quit_gui():
    global root
    root.destroy()
    exit(1)


def main():
    table = get_table()
    parse_balances(table)
    save_table(table)


if __name__ == '__main__':
    root = Tk()
    setup_gui()
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    worksheet = get_worksheet()
    main()
